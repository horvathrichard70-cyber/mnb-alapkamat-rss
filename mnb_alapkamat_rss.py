"""
MNB jegybanki alapkamat -> RSS feed generator

Forrás adat: https://www.mnb.hu/root/BaseRate/BaseRateExcel/alapkamat.xlsx
Forrás oldal: https://www.mnb.hu/Jegybanki_alapkamat_alakulasa

Fájlszerkezet (ellenőrizve a valós alapkamat.xlsx alapján):
    1 munkalap ("Alapkamat"), fejléc az 1. sorban.
    A oszlop: dátum (datetime) - a rendelet hatálybalépésének napja
    B oszlop: kamat szövegként, pl. "6,00%"

Használat:
    python mnb_alapkamat_rss.py            # legenerálja a rss.xml-t, frissíti a last_rate.txt-t
    python mnb_alapkamat_rss.py --inspect  # kiírja a nyers sorokat (hibakereséshez)

Kilépési kód:
    0 -> nem változott az alapkamat az előző futtatáshoz képest
    2 -> változott (ezt használja a workflow az e-mail-küldés feltételeként)

Függőségek: pip install requests openpyxl
"""
import io
import re
import sys
from datetime import datetime, timezone
from email.utils import format_datetime

import requests
import openpyxl

XLSX_URL = "https://www.mnb.hu/root/BaseRate/BaseRateExcel/alapkamat.xlsx"
FEED_TITLE = "MNB jegybanki alapkamat"
FEED_LINK = "https://www.mnb.hu/Jegybanki_alapkamat_alakulasa"
FEED_DESC = "A Magyar Nemzeti Bank jegybanki alapkamatának alakulása"
OUTPUT_FILE = "rss.xml"
STATE_FILE = "last_rate.txt"
MAX_ITEMS = 30

RATE_RE = re.compile(r"(-?\d+(?:[.,]\d+)?)\s*%")


def parse_rate(s: str):
    """'6,00%' -> 6.0"""
    if not isinstance(s, str):
        return None
    m = RATE_RE.search(s)
    if not m:
        return None
    return float(m.group(1).replace(",", "."))


def fetch_workbook() -> openpyxl.Workbook:
    r = requests.get(XLSX_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)


def extract_series(wb: openpyxl.Workbook):
    """Az 'Alapkamat' munkalap A/B oszlopaiból (dátum, kamat%) párokat olvas ki.
    Fejléc sort kihagyja. Csökkenő időrendben adja vissza."""
    ws = wb.worksheets[0]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        date_val, rate_raw = row[0], row[1]
        if not isinstance(date_val, datetime):
            continue
        rate = parse_rate(rate_raw)
        if rate is None:
            continue
        results.append((date_val, rate))
    results.sort(key=lambda x: x[0], reverse=True)
    return results


def build_rss(series) -> str:
    items = []
    for date_val, rate in series[:MAX_ITEMS]:
        pub_date = format_datetime(date_val.replace(tzinfo=timezone.utc))
        rate_str = f"{rate:.2f}".replace(".", ",")
        title = f"Jegybanki alapkamat: {rate_str}%"
        guid = f"mnb-alapkamat-{date_val.strftime('%Y-%m-%d')}"
        items.append(
            f"""    <item>
      <title>{title}</title>
      <link>{FEED_LINK}</link>
      <guid isPermaLink="false">{guid}</guid>
      <pubDate>{pub_date}</pubDate>
      <description>{title}, hatályos: {date_val.strftime('%Y.%m.%d.')}</description>
    </item>"""
        )
    now = format_datetime(datetime.now(timezone.utc))
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
  <channel>
    <title>{FEED_TITLE}</title>
    <link>{FEED_LINK}</link>
    <description>{FEED_DESC}</description>
    <lastBuildDate>{now}</lastBuildDate>
{chr(10).join(items)}
  </channel>
</rss>
"""


def read_previous_rate():
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            content = f.read().strip()
            return content or None
    except FileNotFoundError:
        return None


def write_current_rate(date_val, rate):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        f.write(f"{date_val.strftime('%Y-%m-%d')}|{rate:.2f}")


def inspect(wb: openpyxl.Workbook):
    ws = wb.worksheets[0]
    print(f"--- Munkalap: {ws.title} ({ws.dimensions}) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 15:
            print("... (levágva)")
            break
        print(row)


def main():
    wb = fetch_workbook()

    if "--inspect" in sys.argv:
        inspect(wb)
        return

    series = extract_series(wb)
    if not series:
        print(
            "Nem sikerült adatot kinyerni az XLSX-ből. Futtasd '--inspect' móddal.",
            file=sys.stderr,
        )
        sys.exit(1)

    rss = build_rss(series)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(rss)

    latest_date, latest_rate = series[0]
    current_key = f"{latest_date.strftime('%Y-%m-%d')}|{latest_rate:.2f}"
    previous_key = read_previous_rate()
    write_current_rate(latest_date, latest_rate)

    print(f"Legutóbbi adat: {latest_date.date()} - {latest_rate}%")

    if previous_key is None:
        print("Nincs korábbi állapot (első futás) - nem küldünk értesítést.")
        sys.exit(0)

    if current_key != previous_key:
        print(f"VÁLTOZÁS: {previous_key} -> {current_key}")
        sys.exit(2)

    print("Nincs változás.")
    sys.exit(0)


if __name__ == "__main__":
    main()
